#!/usr/bin/env python3
"""
Audit a voice-agent import file BEFORE the campaign dials.

Every bad row is a call that costs money and returns nothing: a malformed
number rings nowhere, a duplicate debtor gets called twice and complains, an
impossible amount makes the agent say something wrong to a real person.

Design principle (docs/10): deterministic code does the deterministic work.
Phone formatting, arithmetic and exact-key deduplication are Python -- they must
be right every time, and a model would only make them occasionally wrong. The
LLM is used for exactly one thing: deciding whether two similar-looking names at
the same unit are the same person. That is genuinely fuzzy and rules do it badly.

  ./import_audit.py --input import.xlsx
  ./import_audit.py --input import.csv --report audit.md --clean cleaned.csv
  ./import_audit.py --input import.xlsx --no-llm        # rules only, no cluster

Exit codes:  0 clean   1 warnings   2 blocking errors found
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path

# South African numbering. Adjust COUNTRY_CODE / NATIONAL_LEN for other regions.
COUNTRY_CODE = "27"
NATIONAL_LEN = 9          # digits after the country code
MOBILE_PREFIXES = ("6", "7", "8")   # SA mobile ranges, after the leading 0
# Landlines are dialable -- just less likely to reach a person on a collections
# call -- so they are a warning, not a blocking error.
LANDLINE_PREFIXES = ("1", "2", "3", "4", "5")

MAX_SANE_AMOUNT = 10_000_000       # R10m on a levy account is a data error
PLACEHOLDER_NAMES = {"n/a", "na", "unknown", "test", "tbc", "-", "none", "owner",
                     "the owner", "tenant", "resident", "xxx"}


# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------

def load_rows(path: Path) -> tuple[list[dict], list[str]]:
    """Return (rows, columns). Supports .xlsx and .csv/.tsv."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl needed for .xlsx: pip install openpyxl\n"
                     "(or export the sheet to CSV and pass that)")
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(it)]
        rows = []
        for raw in it:
            if all(c is None or str(c).strip() == "" for c in raw):
                continue
            rows.append({h: raw[i] if i < len(raw) else None
                         for i, h in enumerate(header) if h})
        return rows, [h for h in header if h]

    import csv
    delim = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open(newline="", encoding="utf-8-sig") as f:
        rdr = csv.DictReader(f, delimiter=delim)
        return [dict(r) for r in rdr], list(rdr.fieldnames or [])


def find_col(columns: list[str], *candidates: str) -> str | None:
    """Locate a column by fuzzy name match -- source files never agree on headers."""
    low = {c.lower().replace("_", " ").replace("-", " ").strip(): c for c in columns}
    for want in candidates:
        w = want.lower()
        for key, orig in low.items():
            if key == w:
                return orig
    for want in candidates:
        w = want.lower()
        for key, orig in low.items():
            if w in key:
                return orig
    return None


# --------------------------------------------------------------------------
# Deterministic checks
# --------------------------------------------------------------------------

def normalise_phone(raw) -> tuple[str | None, str | None, str | None]:
    """
    Return (e164, error, warning). Deterministic -- never ask a model to do this.

    error   = do not dial this row
    warning = dialable, but worth a look (e.g. a landline on a mobile campaign)
    """
    if raw is None or str(raw).strip() == "":
        return None, "missing", None
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return None, "no digits", None

    if digits.startswith("00" + COUNTRY_CODE):
        digits = digits[2:]

    if digits.startswith(COUNTRY_CODE) and len(digits) == len(COUNTRY_CODE) + NATIONAL_LEN:
        national = digits[len(COUNTRY_CODE):]
    elif digits.startswith("0") and len(digits) == NATIONAL_LEN + 1:
        national = digits[1:]
    elif digits.startswith("0"):
        # A national number that still carries its trunk 0 but is the wrong
        # length is a truncated or over-typed entry, not a strange prefix.
        return None, (f"looks truncated -- {len(digits)} digits starting with 0, "
                      f"expected {NATIONAL_LEN + 1}"), None
    elif len(digits) == NATIONAL_LEN:
        national = digits
    else:
        return None, f"wrong length ({len(digits)} digits)", None

    if len(set(national)) == 1:
        return None, "all-same digits", None

    e164 = f"+{COUNTRY_CODE}{national}"
    if national[0] in MOBILE_PREFIXES:
        return e164, None, None
    if national[0] in LANDLINE_PREFIXES:
        return e164, None, f"landline (0{national[:2]}) -- dialable, lower reach rate"
    return None, f"implausible prefix (0{national[0]})", None


def to_amount(raw) -> float | None:
    if raw is None or str(raw).strip() == "":
        return None
    s = str(raw).strip().replace("R", "").replace(" ", "").replace(",", "")
    neg = s.startswith("(") and s.endswith(")")   # accounting negatives
    if neg:
        s = s[1:-1]
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def name_key(name: str) -> str:
    """Normalise a name for duplicate grouping: drop titles, order-independent."""
    s = re.sub(r"[^a-z ]", " ", str(name).lower())
    parts = [p for p in s.split()
             if p not in {"mr", "mrs", "ms", "miss", "dr", "prof", "the", "and"}]
    return " ".join(sorted(parts))


# --------------------------------------------------------------------------
# The one fuzzy judgement, delegated to the model
# --------------------------------------------------------------------------

def ask_same_person(base_url, api_key, model, a: str, b: str, unit: str,
                    timeout: float = 120) -> tuple[bool, str]:
    prompt = (
        "Two records on the same property unit. Are they the same person?\n\n"
        f"Unit: {unit}\nName A: {a}\nName B: {b}\n\n"
        "Consider initials vs full names, married/maiden names, spelling "
        "variants, and reversed name order. Two different people can share a "
        "surname and live at one unit -- spouses, parent and child.\n\n"
        'Reply with JSON only: {"same": true|false, "why": "<10 words>"}'
    )
    body = json.dumps({
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.0, "max_tokens": 80,
    }).encode()
    req = urllib.request.Request(
        base_url.rstrip("/") + "/v1/chat/completions", data=body,
        headers={"Content-Type": "application/json",
                 **({"Authorization": f"Bearer {api_key}"} if api_key else {})},
        method="POST")
    with urllib.request.urlopen(req, timeout=timeout) as r:
        text = json.loads(r.read())["choices"][0]["message"]["content"]
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m:
        raise ValueError("no JSON in model reply")
    d = json.loads(m.group(0))
    return bool(d.get("same")), str(d.get("why", ""))[:60]


# --------------------------------------------------------------------------
# Audit
# --------------------------------------------------------------------------

def audit(rows, cols, args):
    c_name  = find_col(cols, "name", "customer name", "debtor", "owner")
    c_phone = find_col(cols, "phone", "cell", "mobile", "contact number", "msisdn")
    c_unit  = find_col(cols, "unit reference", "unit", "unit_reference", "account", "erf")
    c_arr   = find_col(cols, "arrears amount", "arrears", "arrears_amount", "overdue")
    c_bal   = find_col(cols, "total balance", "balance", "total_balance")
    c_months= find_col(cols, "months in arrears", "months", "months_in_arrears", "age")

    errors, warnings = [], []
    seen_phone = defaultdict(list)
    unit_names = defaultdict(list)
    clean_rows = []

    for i, row in enumerate(rows, start=2):   # +2: header row, 1-indexed
        name = str(row.get(c_name) or "").strip() if c_name else ""
        unit = str(row.get(c_unit) or "").strip() if c_unit else ""

        # --- phone ---
        if c_phone:
            e164, problem, note = normalise_phone(row.get(c_phone))
            if problem:
                errors.append((i, "phone", f"{problem}: {row.get(c_phone)!r}"))
            else:
                row["_e164"] = e164
                seen_phone[e164].append((i, name))
                if note:
                    warnings.append((i, "phone", note))
        else:
            errors.append((i, "phone", "no phone column found in this file"))

        # --- name ---
        if not name:
            errors.append((i, "name", "missing"))
        elif name.lower() in PLACEHOLDER_NAMES:
            errors.append((i, "name", f"placeholder value {name!r}"))
        elif len(name) < 2:
            warnings.append((i, "name", f"suspiciously short: {name!r}"))
        elif not re.search(r"[A-Za-z]", name):
            errors.append((i, "name", f"no letters: {name!r}"))
        if name and unit:
            unit_names[unit].append((i, name))

        # --- amounts ---
        arrears = to_amount(row.get(c_arr)) if c_arr else None
        balance = to_amount(row.get(c_bal)) if c_bal else None
        if c_arr and arrears is None:
            errors.append((i, "arrears", f"unparseable: {row.get(c_arr)!r}"))
        elif arrears is not None:
            if arrears < 0:
                errors.append((i, "arrears", f"negative ({arrears:,.2f}) -- in credit, do not call"))
            elif arrears == 0:
                warnings.append((i, "arrears", "zero -- nothing to collect"))
            elif arrears > MAX_SANE_AMOUNT:
                errors.append((i, "arrears", f"implausible ({arrears:,.2f})"))
        if arrears is not None and balance is not None and arrears > balance + 0.01:
            errors.append((i, "amounts", f"arrears {arrears:,.2f} exceeds balance {balance:,.2f}"))

        # --- months ---
        if c_months:
            m = to_amount(row.get(c_months))
            if m is None and str(row.get(c_months) or "").strip():
                warnings.append((i, "months", f"unparseable: {row.get(c_months)!r}"))
            elif m is not None and (m < 0 or m > 120):
                warnings.append((i, "months", f"out of range ({m:g})"))

        if not any(e[0] == i for e in errors):
            clean_rows.append(row)

    # --- duplicates: exact phone (deterministic) ---
    dupes = []
    for phone, entries in seen_phone.items():
        if len(entries) > 1:
            dupes.append((phone, entries))

    # --- duplicates: same unit, similar names (model decides) ---
    fuzzy = []
    if not args.no_llm:
        checked = 0
        for unit, entries in unit_names.items():
            if len(entries) < 2 or checked >= args.max_llm_checks:
                continue
            for x in range(len(entries)):
                for y in range(x + 1, len(entries)):
                    (i1, n1), (i2, n2) = entries[x], entries[y]
                    if name_key(n1) == name_key(n2):
                        fuzzy.append((unit, i1, n1, i2, n2, True, "identical after normalising"))
                        continue
                    # Only bother the model when the names share something.
                    if not (set(name_key(n1).split()) & set(name_key(n2).split())):
                        continue
                    if checked >= args.max_llm_checks:
                        break
                    checked += 1
                    try:
                        same, why = ask_same_person(args.base_url, args.api_key,
                                                    args.model, n1, n2, unit)
                        if same:
                            fuzzy.append((unit, i1, n1, i2, n2, True, why))
                    except Exception as e:  # noqa: BLE001
                        warnings.append((i1, "dedupe", f"model check failed: {type(e).__name__}"))
        if checked:
            print(f"  (asked the model about {checked} name pairs)", file=sys.stderr)

    return {
        "columns": {"name": c_name, "phone": c_phone, "unit": c_unit,
                    "arrears": c_arr, "balance": c_bal, "months": c_months},
        "total": len(rows), "errors": errors, "warnings": warnings,
        "phone_dupes": dupes, "fuzzy_dupes": fuzzy, "clean_rows": clean_rows,
    }


def render(res, path: Path) -> str:
    L = [f"# Import audit — {path.name}", ""]
    cols = res["columns"]
    missing = [k for k, v in cols.items() if v is None]
    L += [f"**{res['total']} rows.** Columns detected: " +
          ", ".join(f"`{k}`→`{v}`" for k, v in cols.items() if v) or "none"]
    if missing:
        L.append(f"\n> Not found: {', '.join(missing)}. Checks for those were skipped — "
                 "if the file does have them under another name, rename and re-run.")
    L.append("")

    bad_rows = {e[0] for e in res["errors"]}
    L += ["## Summary", "",
          f"- **{len(bad_rows)} rows blocked** ({len(res['errors'])} errors)",
          f"- {len(res['warnings'])} warnings",
          f"- {len(res['phone_dupes'])} duplicate phone numbers",
          f"- {len(res['fuzzy_dupes'])} likely duplicate people",
          f"- **{len(res['clean_rows'])} rows safe to dial**", ""]

    if res["errors"]:
        L += ["## Errors — these rows should not be called", "",
              "| row | field | problem |", "|---|---|---|"]
        L += [f"| {r} | {f} | {m} |" for r, f, m in res["errors"][:100]]
        if len(res["errors"]) > 100:
            L.append(f"\n_+{len(res['errors']) - 100} more_")
        L.append("")

    if res["phone_dupes"]:
        L += ["## Duplicate phone numbers", "",
              "Calling the same number twice is how a campaign generates complaints.", "",
              "| number | rows |", "|---|---|"]
        L += [f"| `{p}` | " + ", ".join(f"{i} ({n})" for i, n in e) + " |"
              for p, e in res["phone_dupes"][:50]]
        L.append("")

    if res["fuzzy_dupes"]:
        L += ["## Likely the same person", "",
              "| unit | rows | names | why |", "|---|---|---|---|"]
        L += [f"| {u} | {i1}, {i2} | {n1} / {n2} | {why} |"
              for u, i1, n1, i2, n2, _, why in res["fuzzy_dupes"][:50]]
        L.append("")

    if res["warnings"]:
        L += ["## Warnings — dialable, but check", "",
              "| row | field | note |", "|---|---|---|"]
        L += [f"| {r} | {f} | {m} |" for r, f, m in res["warnings"][:60]]
        L.append("")

    return "\n".join(L)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--report", type=Path, help="write the markdown report here")
    ap.add_argument("--clean", type=Path, help="write error-free rows to this CSV")
    ap.add_argument("--base-url", default=os.environ.get("LLM_BASE_URL", "http://192.168.1.201:4000"))
    ap.add_argument("--api-key", default=os.environ.get("LLM_API_KEY", ""))
    ap.add_argument("--model", default="fast",
                    help="small model is right for this -- it is a judgement call, not an essay")
    ap.add_argument("--no-llm", action="store_true", help="rules only, no cluster needed")
    ap.add_argument("--max-llm-checks", type=int, default=200,
                    help="cap model calls so a huge file cannot run all night")
    args = ap.parse_args()

    if not args.input.exists():
        sys.exit(f"not found: {args.input}")
    rows, cols = load_rows(args.input)
    if not rows:
        sys.exit("no data rows found")

    res = audit(rows, cols, args)
    report = render(res, args.input)

    if args.report:
        args.report.write_text(report)
        print(f"report written to {args.report}")
    else:
        print(report)

    if args.clean:
        import csv
        keys = [c for c in cols] + (["_e164"] if any("_e164" in r for r in res["clean_rows"]) else [])
        with args.clean.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
            w.writeheader()
            w.writerows(res["clean_rows"])
        print(f"{len(res['clean_rows'])} clean rows written to {args.clean}")

    bad = len({e[0] for e in res["errors"]})
    if bad:
        print(f"\n{bad} rows blocked. Each one is a wasted call avoided.", file=sys.stderr)
        return 2
    return 1 if res["warnings"] else 0


if __name__ == "__main__":
    sys.exit(main())
